"""
# v 2.1 加入lmdb
"""

import json
import lmdb
import numpy as np
import openpyxl
import pickle
import rich
import os
import shutil
import pathlib
import uuid
import base64
import cv2
from typing import Union
import zipfile
from datetime import datetime


def generate_uuid():
    return str(uuid.uuid4())

def generate_timestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
########################################################################

def load_json_file(file_path:Union[str , pathlib.PosixPath], data_type="dict"):
    if isinstance(file_path,pathlib.PosixPath):
        file_path = str(file_path)
    if data_type == "dict":
        json_file = dict()
    elif data_type == "list":
        json_file = list()
    else:
        raise ValueError("数据类型不对")
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r') as f:
                json_file = json.load(f)
        except IOError as e:
            rich.print(f"[red]无法打开文件{file_path}：{e}")
        except json.JSONDecodeError as e:
            rich.print(f"[red]解析 JSON 文件时出错{file_path}：{e}")
    else:
        rich.print(f"[yellow]{file_path}文件不存在，正在传入空文件...[/yellow]")
    return json_file

def dump_json_file(json_file, file_path:Union[str , pathlib.PosixPath],if_print = True,if_backup = True,if_backup_delete=False):
    if isinstance(file_path,pathlib.PosixPath):
        file_path = str(file_path)
    backup_path = file_path + ".bak"  # 定义备份文件的路径
    if os.path.exists(file_path) and if_backup:
        shutil.copy(file_path, backup_path)
    before_exist = os.path.exists(file_path)
    try:
        with open(file_path, 'w') as f:
            json.dump(json_file, f, indent=4)
        if before_exist and if_print:
            rich.print(f"[yellow]更新{file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]创建{file_path}[/green]")
    except IOError as e:
        if os.path.exists(backup_path) and if_backup:
            shutil.copy(backup_path, file_path)
            if if_print:
                rich.print(f"[red]文件{file_path}写入失败，已从备份恢复原文件: {e}[/red]")
        else:
            if if_print:
                rich.print(f"[red]文件{file_path}写入失败，且无备份可用：{e}[/red]")
    finally:
        # 清理，删除备份文件
        if if_backup:
            if os.path.exists(backup_path) and if_backup_delete:
                os.remove(backup_path)
            if not os.path.exists(backup_path) and not if_backup_delete : #如果一开始是空的
                shutil.copy(file_path, backup_path)
            
class JsonlProcessor:
    
    def __init__(self, file_path:Union[str , pathlib.PosixPath],
                 if_backup = True,
                 if_print=True
                 ):
        
        self.file_path = file_path if not isinstance(file_path,pathlib.PosixPath) else str(file_path)
        
        self.if_print = if_print
        self.if_backup = if_backup

        self._mode = ""

        self._read_file = None
        self._write_file = None
        self._read_position = 0
        self.lines = 0

    @property
    def bak_file_path(self):
        return str(self.file_path) + ".bak"
    
    def exists(self):
        return os.path.exists(self.file_path)

    def len(self):
        file_length = 0
        if not self.exists():
            return file_length
        if self.lines == 0:
            with open(self.file_path, 'r', encoding='utf-8') as file:
                while file.readline():
                    file_length+=1
            self.lines = file_length
        return self.lines

    def close(self,mode = "rw"):
        # 关闭文件资源
        if "r" in mode:
            if self._write_file:
                self._write_file.close()
                self._write_file = None
        if "w" in mode:
            if self._read_file:
                self._read_file.close()
                self._read_file = None
            self.lines = 0
        

    def reset(self, file_path:Union[str , pathlib.PosixPath]):
        self.close()
        self.file_path = file_path if not isinstance(file_path,pathlib.PosixPath) else str(file_path)


    def load_line(self):
        if not self.exists():
            rich.print(f"[yellow]{self.file_path}文件不存在,返回{None}")
            return None
        if self._mode != "r":
            self.close("r")
        if not self._read_file:
            self._read_file = open(self.file_path, 'r', encoding='utf-8')
        
        self._read_file.seek(self._read_position)
        self._mode = "r"
        try:
            line = self._read_file.readline()
            self._read_position = self._read_file.tell()
            if not line:
                self.close()
                return None
            return json.loads(line.strip())
        except json.JSONDecodeError as e:
            self.close()
            rich.print(f"[red]文件{self.file_path}解析出现错误：{e}")
            return None
        except IOError as e:
            self.close()
            rich.print(f"[red]无法打开文件{self.file_path}：{e}")
            return None
    
    def load_lines(self):
        """获取jsonl中的line，直到结尾"""
        lines = []
        while True:
            line = self.load_line()
            if line ==None:
                break
            lines.append(line)
        return lines
        

    def load_restart(self):
        self.close(mode="r")
        self._read_position = 0
         
    def dump_line(self, data):
        if not isinstance(data,dict) and not isinstance(data,list):
            raise ValueError("数据类型不对")
        if self.len() % 50 == 1 and self.if_backup:
            shutil.copy(self.file_path, self.bak_file_path)
        self._mode = "a"
        if not self._write_file:
            self._write_file = open(self.file_path, 'a', encoding='utf-8')
        try:
            json_line = json.dumps(data)
            self._write_file.write(json_line + '\n')
            self._write_file.flush()
            self.lines += 1  
            return True
        except Exception as e:
            
            if os.path.exists(self.bak_file_path) and self.if_backup:
                shutil.copy(self.bak_file_path, self.file_path)
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，已从备份恢复原文件: {e}[/red]")
            else:
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，且无备份可用：{e}[/red]") 
            return False

    def dump_lines(self,datas):
        if not isinstance(datas,list):
            raise ValueError("数据类型不对")
        if self.if_backup and os.path.exists(self.file_path):
            shutil.copy(self.file_path, self.bak_file_path)
        self._mode = "a"
        if not self._write_file:
            self._write_file = open(self.file_path, 'a', encoding='utf-8')
        try:
            self.len()
            for data in datas:
                json_line = json.dumps(data)
                self._write_file.write(json_line + '\n')
                self.lines += 1  
            self._write_file.flush()
            return True
        except Exception as e:
            if os.path.exists(self.bak_file_path) and self.if_backup:
                shutil.copy(self.bak_file_path, self.file_path)
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，已从备份恢复原文件: {e}[/red]")
            else:
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，且无备份可用：{e}[/red]") 
                return False
            
    def dump_restart(self):
        self.close()
        self._mode= "w"
        with open(self.file_path, 'w', encoding='utf-8') as file:
            pass 
          
    def load(self):
        jsonl_file = []
        if self.exists():
            try:
                with open(self.file_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        jsonl_file.append(json.loads(line))
            except IOError as e:
                rich.print(f"[red]无法打开文件：{e}")
            except json.JSONDecodeError as e:
                rich.print(f"[red]解析 JSON 文件时出错：{e}")
        else:
            rich.print(f"[yellow]{self.file_path}文件不存在，正在传入空文件...[/yellow]")
        return jsonl_file

    def dump(self,jsonl_file:list):
        before_exist = self.exists()
        if self.if_backup and before_exist:
            shutil.copy(self.file_path, self.bak_file_path)
        try:
            self.close()
            self._mode = "w"
            with open(self.file_path, 'w', encoding='utf-8') as f:
                for entry in jsonl_file:
                    json_str = json.dumps(entry)
                    f.write(json_str + '\n') 
                    self.lines += 1
            if before_exist and self.if_print:
                rich.print(f"[yellow]更新{self.file_path}[/yellow]")
            elif self.if_print:
                rich.print(f"[green]创建{self.file_path}[/green]")
            return True
        except Exception as e:
            if os.path.exists(self.bak_file_path) and self.if_backup:
                shutil.copy(self.bak_file_path, self.file_path)
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，已从备份恢复原文件: {e}[/red]")
            else:
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，且无备份可用：{e}[/red]") 
            return False
       
def load_jsonl(file_path:Union[str , pathlib.PosixPath]):
    if isinstance(file_path,pathlib.PosixPath):
        file_path = str(file_path)
    jsonl_file = []
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r') as f:
                for line in f:
                    jsonl_file.append(json.loads(line))
        except IOError as e:
            rich.print(f"[red]无法打开文件：{e}")
        except json.JSONDecodeError as e:
            rich.print(f"[red]解析 JSON 文件时出错：{e}")
    else:
        rich.print(f"[yellow]{file_path}文件不存在，正在传入空文件...[/yellow]")
    return jsonl_file

def dump_jsonl(jsonl_file:list,file_path:Union[str , pathlib.PosixPath],if_print=True):
    if isinstance(file_path,pathlib.PosixPath):
        file_path = str(file_path)
    before_exist = os.path.exists(file_path)
    try:
        with open(file_path, 'w') as f:
            for entry in jsonl_file:
                json_str = json.dumps(entry)
                f.write(json_str + '\n') 
        if before_exist and if_print:
            rich.print(f"[yellow]更新{file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]创建{file_path}[/green]")
    except IOError as e:
        print(f"[red]文件{file_path}写入失败，{e}[/red]")   

def load_npy_file(file_path:Union[str , pathlib.PosixPath]):
    if isinstance(file_path,pathlib.PosixPath):
        file_path = str(file_path)
    npy_array = np.empty((0,))
    if os.path.exists(file_path):
        try:
            npy_array = np.load(file_path)
        except IOError as e:
            rich.print(f"[red]无法打开文件：{e}[/red]")
    else:
         rich.print(f"[yellow]{file_path}文件不存在，传入np.empty((0,))[/yellow]")

    return npy_array

def dump_npy_file(npy_array:np.ndarray, file_path:Union[str , pathlib.PosixPath],if_print = True):
    if isinstance(file_path,pathlib.PosixPath):
        file_path = str(file_path)
    before_exist = os.path.exists(file_path)
    try:
        np.save(file_path,npy_array)
        if before_exist and if_print:
            rich.print(f"[yellow]更新{file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]创建{file_path}[/green]")
    except IOError as e:
        rich.print(f"[red]文件写入失败：{e}[/red]")

def load_pickle_file(file_path:Union[str , pathlib.PosixPath]):
    if isinstance(file_path,pathlib.PosixPath):
        file_path = str(file_path)
    pkl_file = {}
    if os.path.exists(file_path):
        try:
            with open(file_path, 'rb') as file:
                # 使用pickle.load加载并反序列化数据
                pkl_file = pickle.load(file)
        except IOError as e:
            rich.print(f"[red]无法打开文件：{e}[/red]")
    else:
         rich.print(f"[yellow]{file_path}文件不存在，传入空文件[/yellow]")

    return pkl_file

def dump_pickle_file(pkl_file, file_path:Union[str , pathlib.PosixPath],if_print = True):
    if isinstance(file_path,pathlib.PosixPath):
        file_path = str(file_path)
    before_exist = os.path.exists(file_path)
    try:
        with open(file_path, 'wb') as file:
            # 使用pickle.dump将数据序列化到文件
            pickle.dump(pkl_file, file)
        if before_exist and if_print:
            rich.print(f"[yellow]更新{file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]创建{file_path}[/green]")
    except IOError as e:
        rich.print(f"[red]文件写入失败：{e}[/red]")

def load_txt_file(file_path:Union[str , pathlib.PosixPath]):
    if isinstance(file_path,pathlib.PosixPath):
        file_path = str(file_path)
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r') as f:
                txt_file = f.read()
        except IOError as e:
            rich.print(f"[red]无法打开文件：{e}[/red]")
    else:
         rich.print(f"[yellow]{file_path}文件不存在，传入空文件[/yellow]")

    return txt_file

def dump_txt_file(file,file_path:Union[str , pathlib.PosixPath],if_print = True):
    if isinstance(file_path,pathlib.PosixPath):
        file_path = str(file_path)
    before_exist = os.path.exists(file_path)
    try:
        with open(file_path, 'w') as f:
            # 使用pickle.dump将数据序列化到文件
            f.write(str(file))
        if before_exist and if_print:
            rich.print(f"[yellow]更新{file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]创建{file_path}[/green]")
    except IOError as e:
        rich.print(f"[red]文件写入失败：{e}[/red]")

def load_excel_file_to_dict(file_path:Union[str , pathlib.PosixPath],if_print = True):
    """存储成如下格式：
    {
        "sheet_name1":[
            {
                "column1":"",
                "column2":"",
                "column3":"",
    }]}
    """
    if isinstance(file_path,str):
        file_path = pathlib.PosixPath(file_path)
    assert file_path.suffix == ".xlsx"
    wb = openpyxl.load_workbook(file_path)
    data = {}
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        sheet_data = []

        for row in rows[1:]:
            row_data = {headers[i]: row[i] for i in range(len(headers))}
            sheet_data.append(row_data)

        data[sheet] = sheet_data
    return data

def dump_excel_file(file:dict, file_path:Union[str , pathlib.PosixPath],if_print = True):
    """转换各种模式为xlsx(excel模式)"""
    if isinstance(file_path,str):
        file_path = pathlib.PosixPath(file_path)
    assert file_path.suffix == ".xlsx"
    
    wb = openpyxl.Workbook()
    
    if isinstance(file, dict):
        """
        如果是dict，暂时要求符合如下格式：
        {
            "sheet_name1":[
                {
                    "column1":"",
                    "column2":"",
                    "column3":"",
        }]}
        """

        sheet0 = list(file.values())[0]
        assert isinstance(sheet0, list)
        row0 = sheet0[0]
        assert isinstance(row0,dict)
        item0 = list(row0.values())[0]
        assert isinstance(item0,str)
        # 然后转成DataFrame模式
        wb.remove(wb.active)  # 移除默认创建的空白工作表
        # 遍历 JSON 数据中的每个工作表
        for sheet_name, rows in file.items():
            ws = wb.create_sheet(title=sheet_name)  # 创建新的工作表
            headers = rows[0].keys()  # 假设所有行的键相同，作为表头
            ws.append(list(headers))  # 添加表头
            for row in rows:
                ws.append(list(row.values()))  # 添加数据行
    try:
        wb.save(file_path)
    except IOError as e:
        rich.print(f"[red]文件写入失败：{e}[/red]")
        
    if file_path.exists() and if_print:
        rich.print(f"[yellow]更新{file_path}[/yellow]")
    elif if_print:
        rich.print(f"[green]创建{file_path}[/green]")

##############################################
    
def zip_fold(source_path:Union[str , pathlib.PosixPath], zip_path:Union[str , pathlib.PosixPath]):
    if isinstance(source_path,str):
        source_path = pathlib.Path(source_path)
    if isinstance(zip_path,str):
        zip_path = pathlib.Path(zip_path)
    if not zip_path.exists():
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(source_path):
                for file in files:
                    # 创建ZIP文件中的文件路径，包括其在文件夹中的相对路径
                    zipf.write(os.path.join(root, file),
                            os.path.relpath(os.path.join(root, file), 
                                            os.path.join(source_path, '..')))
        print(f"[red]{zip_path}已经创建")

def unzip_fold(zip_path:Union[str , pathlib.PosixPath],target_fold:Union[str , pathlib.PosixPath]=None):
    if isinstance(zip_path,str):
        zip_path = pathlib.Path(zip_path)
    if type(target_fold) == type(None):
        parent_path = zip_path.parent
        file_name = zip_path.stem
        target_fold = parent_path / file_name
        pass
    elif isinstance(target_fold,str):   
        target_fold = pathlib.Path(target_fold)
    
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(target_fold)

    print(f"[red]{zip_path}解压到{target_fold}")

def rm_folder(folder_path:Union[str , pathlib.PosixPath]):
    if isinstance(folder_path,str):
        folder_path = pathlib.Path(folder_path)
    if folder_path.exists() and folder_path.is_dir():
        shutil.rmtree(folder_path)
        print(f"Folder '{folder_path}' and its contents have been deleted.")
    else:
        print(f"Folder '{folder_path}' does not exist or is not a directory.")

################################################

class LmdbProcessor:
    def __init__(self,path:Union[str , pathlib.PosixPath],map_size = 10485760): #10mb
        self.path = str(path)
        self.map_size = map_size
        self.env = lmdb.open(self.path,map_size=map_size)
    
    def insert(self,key:str,value):
        key = key.encode()
        value = pickle.dumps(value)
        with self.env.begin(write=True) as txn:
            txn.put(key, value)
        
    def delete(self,key:str):
        key = key.encode()
        try:
            with self.env.begin(write=True) as txn:
                txn.delete(key)
            return True
        except:
            return False
            
    def get(self,key:str):
        key = key.encode()
        with self.env.begin() as txn:
            # 读取数据
            value = txn.get(key)
        if value is not None:
            value = pickle.loads(value)
        return value
    
    def get_all_keys(self):
            # 创建一个游标来遍历数据库
        with self.env.begin() as txn:
            cursor = txn.cursor()
            
            # 遍历所有的键
            keys = []
            for key, _ in cursor:
                keys.append(key)
        return keys
        
    def get_info(self):
        with self.env.begin() as txn:
            cursor = txn.cursor()
            
            # 遍历所有的键
            info = {}
            for key, value in cursor:
                key = key.decode('utf-8')
                if value is not None:
                    value = pickle.loads(value)
                    info[key] = value
        return info
        
    def close(self):
        self.env.close()
        

################################################

def encode_image_to_base64(image:Union[str , pathlib.PosixPath, np.ndarray]):
    """将数据处理为base64 """
    if isinstance(image, str):
        image = pathlib.Path(image)
    if isinstance(image,np.ndarray):
        result = base64.b64encode(image).decode('utf-8')
        return result
    with image.open('rb') as image_file:
        # 对图片数据进行base64编码，并解码为utf-8字符串
        result = base64.b64encode(image_file.read()).decode('utf-8')
        
    return result

################################################

if __name__ == "__main__":
    jp = JsonlProcessor("temp/1.jsonl")
    jp.dump_restart()
    list1 = [a for a in range(10)]
    jp.dump_lines(list1)
    lines = jp.load_lines()
    print(lines)
    list2 = [a for a in range(10,20)]
    jp.dump_lines(list2)
    lines = jp.load_lines()
    print(lines)
    jp.dump_lines(list2)
    lines = jp.load_lines()
    print(lines)
    jp.close()